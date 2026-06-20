"""Importazione dei cronoprogrammi inviati dalle imprese (Excel o PDF).

Ogni impresa invia al project manager un proprio cronoprogramma con l'elenco
delle attività (lavorazioni) e le relative date/importi/%. Questo modulo:

  * legge il file (.xlsx/.xls oppure .pdf);
  * riconosce in modo flessibile le colonne dall'intestazione
    (es. "Attività", "Data inizio", "Data fine", "Importo", "Avanzamento %");
  * legge i metadati (Cantiere, Impresa, Lotto, Categoria, Responsabile)
    da celle etichettate tipo "Cantiere: ...";
  * restituisce un oggetto Lotto pronto da inserire nell'archivio.

Il formato non deve essere perfetto: il parser cerca le parole chiave nelle
intestazioni, quindi tollera piccole variazioni di nomi e posizioni.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from .models import Attivita, Lotto

# parole chiave per riconoscere le colonne della tabella attività
_COLONNE = {
    "nome": ["attiv", "descriz", "lavorazion", "voce", "fase"],
    "data_inizio": ["inizio", "start", "da"],
    "data_fine": ["fine", "termine", "end", "ultim"],
    "importo": ["importo", "valore", "euro", "€", "prezzo", "computo"],
    "avanzamento": ["avanz", "compl", "%", "stato", "eseguito"],
}

# etichette dei metadati cercate nelle prime righe del file
_METADATI = {
    "cantiere": ["cantiere", "commessa", "appalto", "opera"],
    "impresa": ["impresa", "operatore", "ditta", "appaltatore", "esecutore"],
    "lotto": ["lotto", "sotto computo", "sottocomputo", "affidamento"],
    "categoria": ["categoria", "tipologia"],
    "responsabile": ["responsabile", "referente", "direttore"],
    "importo_lotto": ["importo lotto", "importo contratto", "valore lotto",
                      "totale affidamento", "importo totale"],
}


@dataclass
class CronoprogrammaImportato:
    """Risultato dell'import di un singolo file."""

    cantiere_nome: str
    lotto: Lotto
    file_origine: str
    righe_lette: int


# ---------------------------------------------------------- parsing di valori

def parse_numero(valore) -> float:
    """Converte testo/numero in float, gestendo formati € e separatori IT.

    Esempi accettati: 1234.56, "1.234,56", "€ 12.000", "45%".
    """
    if valore is None:
        return 0.0
    if isinstance(valore, (int, float)):
        return float(valore)
    testo = str(valore)
    testo = re.sub(r"[^0-9,.\-]", "", testo)  # via €, %, spazi, lettere
    if not testo:
        return 0.0
    # se ci sono sia '.' che ',', l'ultimo è il separatore decimale
    if "," in testo and "." in testo:
        if testo.rfind(",") > testo.rfind("."):
            testo = testo.replace(".", "").replace(",", ".")
        else:
            testo = testo.replace(",", "")
    elif "," in testo:
        testo = testo.replace(",", ".")
    try:
        return float(testo)
    except ValueError:
        return 0.0


def parse_data(valore) -> date | None:
    """Converte testo/datetime in date. Gestisce gg/mm/aaaa, aaaa-mm-gg, ecc."""
    if valore is None or valore == "":
        return None
    if isinstance(valore, datetime):
        return valore.date()
    if isinstance(valore, date):
        return valore
    testo = str(valore).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(testo, fmt).date()
        except ValueError:
            continue
    return None


def parse_percentuale(valore) -> float | None:
    """Converte una % in float 0-100. Restituisce None se la cella è vuota."""
    if valore is None or str(valore).strip() == "":
        return None
    num = parse_numero(valore)
    # se espressa come frazione (0-1), riportala a 0-100
    if 0 < num <= 1 and "%" not in str(valore):
        num *= 100
    return round(num, 1)


# --------------------------------------------------- lettura righe dai file

def _righe_da_excel(percorso: Path) -> list[list]:
    """Legge tutte le righe del primo foglio come liste di valori."""
    from openpyxl import load_workbook

    wb = load_workbook(percorso, data_only=True)
    ws = wb.active
    return [list(riga) for riga in ws.iter_rows(values_only=True)]


def _soglia_colonne(parole: list[dict]) -> float:
    """Stima la distanza orizzontale che separa due colonne in una riga.

    Calcola gli spazi fra parole consecutive e cerca il "salto" più grande
    fra spazi piccoli (dentro la stessa colonna) e spazi grandi (fra colonne).
    """
    parole = sorted(parole, key=lambda w: w["x0"])
    gap = [b["x0"] - a["x1"] for a, b in zip(parole, parole[1:]) if b["x0"] > a["x1"]]
    if not gap:
        return 9999.0
    gap_ord = sorted(gap)
    if len(gap_ord) == 1:
        return gap_ord[0] / 2
    # salto massimo fra spazi ordinati: confine fra "piccoli" e "grandi"
    salti = [(gap_ord[i + 1] - gap_ord[i], i) for i in range(len(gap_ord) - 1)]
    _, i = max(salti)
    return (gap_ord[i] + gap_ord[i + 1]) / 2


def _confini_colonne(parole: list[dict], soglia: float) -> list[float]:
    """Ritorna le x dei confini fra colonne, ricavate dagli spazi > soglia."""
    parole = sorted(parole, key=lambda w: w["x0"])
    confini = []
    for a, b in zip(parole, parole[1:]):
        if b["x0"] - a["x1"] > soglia:
            confini.append((a["x1"] + b["x0"]) / 2)
    return confini


def _assegna_a_colonne(parole: list[dict], confini: list[float]) -> list[str]:
    """Distribuisce le parole nelle colonne definite dai confini x."""
    celle = [""] * (len(confini) + 1)
    for w in sorted(parole, key=lambda w: w["x0"]):
        centro = (w["x0"] + w["x1"]) / 2
        col = sum(1 for c in confini if centro > c)
        celle[col] = (celle[col] + " " + w["text"]).strip()
    return celle


def _righe_da_pdf(percorso: Path) -> list[list]:
    """Estrae le righe dal PDF ricostruendo le colonne dalla posizione delle parole.

    Strategia:
      1. se la pagina ha tabelle con bordi, usa quelle (celle già separate);
      2. altrimenti raggruppa le parole per riga (coordinata verticale) e
         ricostruisce le colonne dagli spazi orizzontali, usando come
         riferimento i confini della riga di intestazione.

    Ogni riga viene aggiunta sia come testo unico (utile per i metadati) sia,
    nella zona tabellare, come celle separate (utile per le attività).
    """
    import pdfplumber

    righe: list[list] = []
    with pdfplumber.open(percorso) as pdf:
        for pagina in pdf.pages:
            tabelle = pagina.extract_tables() or []
            if tabelle:
                for tabella in tabelle:
                    for riga in tabella:
                        righe.append(list(riga))
                continue

            # nessuna tabella con bordi: ricostruisco dalle parole
            parole = pagina.extract_words() or []
            linee: dict[int, list[dict]] = {}
            for w in parole:
                chiave = round(w["top"] / 3)  # tolleranza verticale ~3pt
                linee.setdefault(chiave, []).append(w)
            linee_ord = [linee[k] for k in sorted(linee)]

            # individua i confini di colonna dalla riga di intestazione
            confini: list[float] = []
            for ln in linee_ord:
                testo = " ".join(w["text"] for w in sorted(ln, key=lambda w: w["x0"]))
                basso = testo.lower()
                if any(p in basso for p in _COLONNE["nome"]) and (
                    any(p in basso for p in _COLONNE["data_inizio"])
                    or any(p in basso for p in _COLONNE["data_fine"])
                ):
                    confini = _confini_colonne(ln, _soglia_colonne(ln))
                    break

            for ln in linee_ord:
                testo = " ".join(w["text"] for w in sorted(ln, key=lambda w: w["x0"]))
                righe.append([testo])  # riga intera (metadati)
                if confini:
                    righe.append(_assegna_a_colonne(ln, confini))  # celle separate
    return righe


# ----------------------------------------------------- riconoscimento tabella

def _testo(cella) -> str:
    return "" if cella is None else str(cella).strip()


def _trova_metadati(righe: list[list]) -> dict[str, str]:
    """Cerca nelle righe le etichette tipo 'Cantiere: ...' e ne estrae il valore."""
    trovati: dict[str, str] = {}
    for riga in righe[:25]:
        celle = [_testo(c) for c in riga]
        for i, cella in enumerate(celle):
            basso = cella.lower()
            for chiave, parole in _METADATI.items():
                if chiave in trovati:
                    continue
                if any(p in basso for p in parole):
                    # valore dopo i ':' nella stessa cella...
                    if ":" in cella:
                        valore = cella.split(":", 1)[1].strip()
                        if valore:
                            trovati[chiave] = valore
                            continue
                    # ...oppure nella prima cella non vuota successiva
                    for succ in celle[i + 1:]:
                        if succ:
                            trovati[chiave] = succ
                            break
    return trovati


def _trova_intestazione(righe: list[list]) -> tuple[int, dict[str, int]]:
    """Individua la riga di intestazione della tabella e mappa colonna->indice.

    Restituisce (indice_riga, {campo: indice_colonna}). Se non trova una
    tabella valida restituisce (-1, {}).
    """
    migliore_riga = -1
    migliore_mappa: dict[str, int] = {}
    migliore_punteggio = 0
    for r, riga in enumerate(righe):
        mappa: dict[str, int] = {}
        for c, cella in enumerate(riga):
            basso = _testo(cella).lower()
            if not basso:
                continue
            for campo, parole in _COLONNE.items():
                if campo in mappa:
                    continue
                if any(p in basso for p in parole):
                    mappa[campo] = c
        # serve almeno il nome attività e una data per considerarla valida
        if "nome" not in mappa or not (
            "data_inizio" in mappa or "data_fine" in mappa
        ):
            continue
        # le colonne devono essere su indici distinti (no riga a cella unica)
        distinti = len(set(mappa.values()))
        if distinti < 2:
            continue
        if distinti > migliore_punteggio:
            migliore_punteggio = distinti
            migliore_mappa = mappa
            migliore_riga = r
    return migliore_riga, migliore_mappa


def _e_riga_totale(nome: str) -> bool:
    return any(k in nome.lower() for k in ("totale", "sommano", "tot.", "subtotale"))


# -------------------------------------------------------------- import file

def importa_file(percorso: str | Path) -> CronoprogrammaImportato:
    """Importa un cronoprogramma da file Excel o PDF e restituisce un Lotto."""
    percorso = Path(percorso)
    if not percorso.exists():
        raise FileNotFoundError(f"File non trovato: {percorso}")

    suffisso = percorso.suffix.lower()
    if suffisso in (".xlsx", ".xlsm", ".xls"):
        righe = _righe_da_excel(percorso)
    elif suffisso == ".pdf":
        righe = _righe_da_pdf(percorso)
    else:
        raise ValueError(f"Formato non supportato: {suffisso} (usa .xlsx o .pdf)")

    meta = _trova_metadati(righe)
    r_header, mappa = _trova_intestazione(righe)
    if r_header < 0:
        raise ValueError(
            "Impossibile riconoscere la tabella delle attività nel file. "
            "Verifica che ci sia una riga di intestazione con almeno "
            "'Attività' e 'Data inizio'/'Data fine'."
        )

    indice_massimo = max(mappa.values())
    attivita: list[Attivita] = []
    for riga in righe[r_header + 1:]:
        # salta le righe "a cella unica" (testo intero) usate solo per i metadati:
        # una riga dati valida ha tante celle quante le colonne della tabella.
        if len(riga) <= indice_massimo:
            continue

        def cella(campo: str):
            i = mappa.get(campo)
            return riga[i] if i is not None and i < len(riga) else None

        nome = _testo(cella("nome"))
        if not nome:
            continue
        if _e_riga_totale(nome):
            continue
        attivita.append(
            Attivita(
                nome=nome,
                data_inizio=parse_data(cella("data_inizio")),
                data_fine=parse_data(cella("data_fine")),
                importo=parse_numero(cella("importo")),
                avanzamento_pct=parse_percentuale(cella("avanzamento")),
            )
        )

    impresa = meta.get("impresa", "Impresa non indicata")
    lotto_nome = meta.get("lotto") or f"Lotto - {impresa}"
    importo_lotto = parse_numero(meta.get("importo_lotto", 0)) or sum(
        a.importo for a in attivita
    )

    lotto = Lotto(
        nome=lotto_nome,
        impresa=impresa,
        categoria=meta.get("categoria", ""),
        responsabile=meta.get("responsabile", ""),
        importo_contratto=importo_lotto,
        attivita=attivita,
        note=f"Importato da {percorso.name}",
    )
    inizio, fine = lotto.date_estremi_attivita()
    lotto.data_inizio = inizio
    lotto.data_fine_prevista = fine

    return CronoprogrammaImportato(
        cantiere_nome=meta.get("cantiere", "Cantiere da definire"),
        lotto=lotto,
        file_origine=percorso.name,
        righe_lette=len(attivita),
    )
