"""Esportazione del cronoprogramma in un file Excel con diagramma di Gantt.

Il file generato contiene, per ogni cantiere:
  * un foglio "Riepilogo" con l'andamento globale e per impresa;
  * un foglio "Gantt" con le barre temporali dei lotti, colorate in base
    al ritardo (verde = in linea, giallo = lieve ritardo, rosso = grave).

Usa la libreria openpyxl (installabile con: pip install openpyxl).
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .analisi import (
    StatoCantiere,
    _avanzamento_atteso_periodo,
    analizza_cantiere,
)
from .models import Archivio, Cantiere

# Colori (esadecimale ARGB) usati nelle barre del Gantt
_VERDE = PatternFill("solid", fgColor="63BE7B")
_GIALLO = PatternFill("solid", fgColor="FFD666")
_ROSSO = PatternFill("solid", fgColor="F8696B")
_GRIGIO = PatternFill("solid", fgColor="D9D9D9")
_INTESTAZIONE = PatternFill("solid", fgColor="305496")
_SOTTOTESTA = PatternFill("solid", fgColor="8EAADB")   # intestazione di lotto
_DA_COMPILARE = PatternFill("solid", fgColor="FFF2CC")  # celle che l'impresa compila
_PRECOMPILATO = PatternFill("solid", fgColor="E2EFDA")  # celle gia' compilate dal PM

_BORDO = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
_BIANCO_BOLD = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)


def _colore_per_scostamento(scostamento: float) -> PatternFill:
    if scostamento < -10:
        return _ROSSO
    if scostamento < -2:
        return _GIALLO
    return _VERDE


def _foglio_riepilogo(wb: Workbook, stato: StatoCantiere) -> None:
    ws = wb.create_sheet(title=_nome_foglio(stato.cantiere.nome, "Riepilogo"))
    c = stato.cantiere

    ws["A1"] = f"CANTIERE: {c.nome}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Tipo: {c.tipo}   Committente: {c.committente}"
    ws["A3"] = (
        f"Avanzamento globale: {stato.avanzamento_globale_reale}% "
        f"(atteso {stato.avanzamento_globale_atteso}%)  "
        f"Scostamento: {stato.scostamento_globale:+}%"
    )
    ws["A3"].font = _BOLD
    ws["A4"] = f"Importo maturato (somma ultimi SAL): € {stato.importo_maturato_totale:,.2f}"

    intestazioni = [
        "Lotto", "Impresa", "Categoria", "Importo €",
        "Avanz. reale %", "Avanz. atteso %", "Scostamento %", "Giudizio",
        "Ultimo SAL €",
    ]
    riga0 = 6
    for col, testo in enumerate(intestazioni, start=1):
        cell = ws.cell(row=riga0, column=col, value=testo)
        cell.fill = _INTESTAZIONE
        cell.font = _BIANCO_BOLD
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _BORDO

    for i, s in enumerate(stato.stati_lotti, start=1):
        r = riga0 + i
        valori = [
            s.lotto.nome, s.lotto.impresa, s.lotto.categoria,
            s.lotto.importo_contratto, s.avanzamento_reale,
            s.avanzamento_atteso, s.scostamento, s.giudizio,
            s.importo_maturato,
        ]
        for col, val in enumerate(valori, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDO
        # colora la cella del giudizio in base al ritardo
        ws.cell(row=r, column=8).fill = _colore_per_scostamento(s.scostamento)

    for col in range(1, len(intestazioni) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _foglio_gantt(wb: Workbook, stato: StatoCantiere) -> None:
    ws = wb.create_sheet(title=_nome_foglio(stato.cantiere.nome, "Gantt"))
    lotti = [s.lotto for s in stato.stati_lotti]
    date_inizio = [l.data_inizio for l in lotti if l.data_inizio]
    date_fine = [l.data_fine_prevista for l in lotti if l.data_fine_prevista]
    if not date_inizio or not date_fine:
        ws["A1"] = "Date non sufficienti per generare il Gantt."
        return

    inizio_min = min(date_inizio)
    fine_max = max(date_fine)
    # scala settimanale per non avere troppe colonne
    giorni_totali = (fine_max - inizio_min).days or 1
    passo = max(1, giorni_totali // 40)  # ~40 colonne max

    col_base = 3  # le prime 2 colonne sono Lotto/Impresa
    ws.cell(row=1, column=1, value="Lotto").font = _BIANCO_BOLD
    ws.cell(row=1, column=2, value="Impresa").font = _BIANCO_BOLD
    ws.cell(row=1, column=1).fill = _INTESTAZIONE
    ws.cell(row=1, column=2).fill = _INTESTAZIONE

    # intestazione temporale
    n_colonne = giorni_totali // passo + 1
    for k in range(n_colonne):
        giorno = inizio_min + timedelta(days=k * passo)
        cell = ws.cell(row=1, column=col_base + k, value=giorno.strftime("%d/%m"))
        cell.fill = _INTESTAZIONE
        cell.font = _BIANCO_BOLD
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_base + k)].width = 4

    oggi = date.today()
    for i, s in enumerate(stato.stati_lotti, start=1):
        r = i + 1
        l = s.lotto
        ws.cell(row=r, column=1, value=l.nome).border = _BORDO
        ws.cell(row=r, column=2, value=l.impresa).border = _BORDO
        if not l.data_inizio or not l.data_fine_prevista:
            continue
        colore = _colore_per_scostamento(s.scostamento)
        for k in range(n_colonne):
            giorno = inizio_min + timedelta(days=k * passo)
            cell = ws.cell(row=r, column=col_base + k)
            if l.data_inizio <= giorno <= l.data_fine_prevista:
                cell.fill = colore
                cell.border = _BORDO

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20

    # legenda
    rl = len(stato.stati_lotti) + 3
    ws.cell(row=rl, column=1, value="Legenda:").font = _BOLD
    for j, (txt, fill) in enumerate(
        [("In linea", _VERDE), ("Ritardo lieve", _GIALLO), ("Ritardo grave", _ROSSO)]
    ):
        c = ws.cell(row=rl + 1 + j, column=1, value=txt)
        ws.cell(row=rl + 1 + j, column=2).fill = fill


def _mesi_tra(inizio: date, fine: date) -> list[tuple[int, int]]:
    """Lista di (anno, mese) dal mese di inizio a quello di fine, inclusi."""
    mesi = []
    a, m = inizio.year, inizio.month
    while (a, m) <= (fine.year, fine.month):
        mesi.append((a, m))
        m += 1
        if m > 12:
            m, a = 1, a + 1
    return mesi


def _foglio_cronoprogramma_generale(wb: Workbook, stato: StatoCantiere,
                                    alla_data: date) -> None:
    """Foglio per la stazione appaltante: tutti i lotti/imprese con il
    dettaglio delle attività e un Gantt mensile colorato per ritardo.
    """
    ws = wb.create_sheet(title=_nome_foglio(stato.cantiere.nome, "Cronoprog"))
    c = stato.cantiere

    ws["A1"] = f"CRONOPROGRAMMA GENERALE - {c.nome}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Committente: {c.committente}   Tipo: {c.tipo}   "
                f"Aggiornato al {alla_data.strftime('%d/%m/%Y')}")
    ws["A3"] = (f"Avanzamento globale: {stato.avanzamento_globale_reale}% "
                f"(atteso {stato.avanzamento_globale_atteso}%)  "
                f"scostamento {stato.scostamento_globale:+}%")
    ws["A3"].font = _BOLD

    # intervallo temporale complessivo (da tutte le attività)
    tutte_inizio, tutte_fine = [], []
    for s in stato.stati_lotti:
        for a in s.lotto.attivita:
            if a.data_inizio:
                tutte_inizio.append(a.data_inizio)
            if a.data_fine:
                tutte_fine.append(a.data_fine)
    # ricade sulle date di lotto se non ci sono attività datate
    for s in stato.stati_lotti:
        if s.lotto.data_inizio:
            tutte_inizio.append(s.lotto.data_inizio)
        if s.lotto.data_fine_prevista:
            tutte_fine.append(s.lotto.data_fine_prevista)
    if not tutte_inizio or not tutte_fine:
        ws["A5"] = "Date non sufficienti per generare il Gantt."
        return
    mesi = _mesi_tra(min(tutte_inizio), max(tutte_fine))

    col_base = 6  # 5 colonne descrittive prima del Gantt
    r_head = 5
    intestazioni = ["Lotto / Attività", "Impresa", "Inizio", "Fine", "Avanz. %"]
    for i, testo in enumerate(intestazioni, start=1):
        cell = ws.cell(row=r_head, column=i, value=testo)
        cell.fill = _INTESTAZIONE
        cell.font = _BIANCO_BOLD
        cell.border = _BORDO
    nomi_mesi = ["gen", "feb", "mar", "apr", "mag", "giu",
                 "lug", "ago", "set", "ott", "nov", "dic"]
    for k, (anno, mese) in enumerate(mesi):
        cell = ws.cell(row=r_head, column=col_base + k,
                       value=f"{nomi_mesi[mese - 1]}\n{str(anno)[2:]}")
        cell.fill = _INTESTAZIONE
        cell.font = _BIANCO_BOLD
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_base + k)].width = 5

    def indice_mese(d: date) -> int:
        return mesi.index((d.year, d.month)) if (d.year, d.month) in mesi else -1

    riga = r_head + 1
    for s in stato.stati_lotti:
        l = s.lotto
        # riga di intestazione del lotto (impresa)
        ws.cell(row=riga, column=1,
                value=f"{l.nome}").font = _BOLD
        for col in range(1, 5):
            ws.cell(row=riga, column=col).fill = _SOTTOTESTA
        ws.cell(row=riga, column=2, value=l.impresa).font = _BOLD
        ws.cell(row=riga, column=5,
                value=f"{s.avanzamento_reale}%").fill = _colore_per_scostamento(
            s.scostamento)
        # barra di sintesi del lotto sull'intera durata
        if l.data_inizio and l.data_fine_prevista:
            i0, i1 = indice_mese(l.data_inizio), indice_mese(l.data_fine_prevista)
            for k in range(i0, i1 + 1):
                if k >= 0:
                    ws.cell(row=riga, column=col_base + k).fill = _SOTTOTESTA
        riga += 1

        # righe di dettaglio: ogni attività del lotto
        for a in l.attivita:
            ws.cell(row=riga, column=1, value=f"   {a.nome}").border = _BORDO
            ws.cell(row=riga, column=2, value=l.impresa).border = _BORDO
            if a.data_inizio:
                ws.cell(row=riga, column=3,
                        value=a.data_inizio.strftime("%d/%m/%y")).border = _BORDO
            if a.data_fine:
                ws.cell(row=riga, column=4,
                        value=a.data_fine.strftime("%d/%m/%y")).border = _BORDO
            atteso = _avanzamento_atteso_periodo(a.data_inizio, a.data_fine, alla_data)
            reale = a.avanzamento_pct if a.avanzamento_pct is not None else atteso
            ws.cell(row=riga, column=5, value=round(reale, 1)).border = _BORDO
            colore = _colore_per_scostamento(round(reale - atteso, 1))
            if a.data_inizio and a.data_fine:
                i0, i1 = indice_mese(a.data_inizio), indice_mese(a.data_fine)
                for k in range(i0, i1 + 1):
                    if k >= 0:
                        cell = ws.cell(row=riga, column=col_base + k)
                        cell.fill = colore
                        cell.border = _BORDO
            riga += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 9

    # legenda
    rl = riga + 2
    ws.cell(row=rl, column=1, value="Legenda colori:").font = _BOLD
    for j, (txt, fill) in enumerate(
        [("In linea", _VERDE), ("Ritardo lieve", _GIALLO), ("Ritardo grave", _ROSSO)]
    ):
        ws.cell(row=rl + 1 + j, column=1, value=txt)
        ws.cell(row=rl + 1 + j, column=2).fill = fill


def _nome_foglio(nome_cantiere: str, suffisso: str) -> str:
    """Excel limita i nomi dei fogli a 31 caratteri e vieta alcuni simboli."""
    pulito = "".join(ch for ch in nome_cantiere if ch not in r'[]:*?/\\')
    base = f"{pulito[:20]} {suffisso}"
    return base[:31]


# Attività standard (WBS di cantiere) precompilate dal PM nel modello.
# Ogni impresa tiene le righe pertinenti al proprio lotto, ne aggiunge se serve
# e compila date / importo / % per ciascuna.
ATTIVITA_STANDARD = [
    "Allestimento cantiere e sicurezza",
    "Scavi e movimenti terra",
    "Opere strutturali / fondazioni",
    "Opere edili e murature",
    "Impianti (elettrico/meccanico/idrico)",
    "Serramenti e facciate",
    "Finiture e tinteggiature",
    "Collaudi, prove e consegna",
]


def genera_modello_cantiere(
    cantiere_nome: str,
    committente: str = "",
    cup: str = "",
    cig: str = "",
    attivita_standard: list[str] | None = None,
    righe_extra: int = 6,
    percorso: str | Path = "export/modello_impresa.xlsx",
) -> Path:
    """Crea il modello Excel PRECOMPILATO da consegnare a ogni impresa.

    Il project manager precompila i riferimenti del cantiere e l'elenco delle
    attività standard (WBS). L'impresa deve solo:
      * indicare la propria Impresa e il Lotto / Commessa (nelle note in alto);
      * compilare per ogni attività: date, importo e % di avanzamento;
      * eventualmente aggiungere righe di attività nelle righe vuote.

    Le celle verdi sono già compilate dal PM, quelle gialle sono da compilare
    a cura dell'impresa. Il formato è quello riconosciuto automaticamente
    dall'import, così tutti i cronoprogrammi ricevuti sono standardizzati.
    """
    attivita_standard = attivita_standard or ATTIVITA_STANDARD
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronoprogramma"

    ws["A1"] = "CRONOPROGRAMMA DI LOTTO - modello da compilare a cura dell'impresa"
    ws["A1"].font = Font(bold=True, size=13, color="305496")

    # blocco riferimenti: alcune righe precompilate dal PM, altre da compilare
    righe_meta = [
        ("Cantiere:", cantiere_nome, _PRECOMPILATO),
        ("Committente:", committente, _PRECOMPILATO),
        ("CUP:", cup, _PRECOMPILATO),
        ("CIG:", cig, _PRECOMPILATO),
        ("Impresa:", "", _DA_COMPILARE),
        ("Lotto / Commessa:", "", _DA_COMPILARE),
        ("Categoria:", "", _DA_COMPILARE),
        ("Responsabile:", "", _DA_COMPILARE),
    ]
    r = 3
    for etichetta, valore, fill in righe_meta:
        ws.cell(row=r, column=1, value=etichetta).font = _BOLD
        cella = ws.cell(row=r, column=2, value=valore)
        cella.fill = fill
        cella.border = _BORDO
        r += 1

    ws.cell(row=r + 1, column=1,
            value="Compilare le celle GIALLE. Le celle VERDI sono già impostate dal PM.")
    ws.cell(row=r + 1, column=1).font = Font(italic=True, color="808080")

    # tabella attività
    r0 = r + 3
    intestazioni = ["Attività", "Data inizio", "Data fine", "Importo €",
                    "Avanzamento %", "Note"]
    for c, testo in enumerate(intestazioni, start=1):
        cell = ws.cell(row=r0, column=c, value=testo)
        cell.fill = _INTESTAZIONE
        cell.font = _BIANCO_BOLD
        cell.border = _BORDO
        cell.alignment = Alignment(horizontal="center")

    n_att = len(attivita_standard)
    for i, nome_att in enumerate(attivita_standard, start=1):
        rr = r0 + i
        # nome attività precompilato (verde), resto da compilare (giallo)
        c_nome = ws.cell(row=rr, column=1, value=nome_att)
        c_nome.fill = _PRECOMPILATO
        c_nome.border = _BORDO
        for c in range(2, len(intestazioni) + 1):
            cc = ws.cell(row=rr, column=c)
            cc.fill = _DA_COMPILARE
            cc.border = _BORDO
    # righe vuote per attività aggiuntive
    for i in range(n_att + 1, n_att + 1 + righe_extra):
        rr = r0 + i
        for c in range(1, len(intestazioni) + 1):
            cc = ws.cell(row=rr, column=c)
            cc.fill = _DA_COMPILARE
            cc.border = _BORDO

    # suggerimento formati
    rsugg = r0 + n_att + righe_extra + 2
    ws.cell(row=rsugg, column=1,
            value="Formato date: AAAA-MM-GG (es. 2026-03-15). "
                  "Avanzamento %: 0-100 (lasciare vuoto se non disponibile).")
    ws.cell(row=rsugg, column=1).font = Font(italic=True, color="808080")

    for c, larg in enumerate([38, 14, 14, 14, 14, 30], start=1):
        ws.column_dimensions[get_column_letter(c)].width = larg

    percorso = Path(percorso)
    percorso.parent.mkdir(parents=True, exist_ok=True)
    wb.save(percorso)
    return percorso


# alias per retrocompatibilità con il menu/precedenti versioni
def genera_template_impresa(percorso: str | Path = "export/modello_impresa.xlsx") -> Path:
    """Genera un modello generico (cantiere da indicare a cura dell'impresa)."""
    return genera_modello_cantiere(cantiere_nome="", percorso=percorso)


def esporta(archivio: Archivio, percorso: str | Path = "export/cronoprogramma.xlsx",
            alla_data: date | None = None) -> Path:
    """Genera il file Excel con un foglio Riepilogo + Gantt per ogni cantiere."""
    wb = Workbook()
    wb.remove(wb.active)  # rimuove il foglio vuoto di default

    if not archivio.cantieri:
        ws = wb.create_sheet("Vuoto")
        ws["A1"] = "Nessun cantiere presente."
    else:
        riferimento = alla_data or date.today()
        for cantiere in archivio.cantieri:
            stato = analizza_cantiere(cantiere, alla_data)
            _foglio_riepilogo(wb, stato)
            _foglio_cronoprogramma_generale(wb, stato, riferimento)
            _foglio_gantt(wb, stato)

    percorso = Path(percorso)
    percorso.parent.mkdir(parents=True, exist_ok=True)
    wb.save(percorso)
    return percorso
