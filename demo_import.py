"""Dimostra l'import e la fusione dei cronoprogrammi delle imprese.

1. Genera 3 file Excel di esempio (come li invierebbero 3 imprese diverse),
   con formati leggermente differenti per provare la flessibilità del parser.
2. Li importa e li unisce in un unico cantiere.
3. Mostra l'analisi (avanzamenti calcolati automaticamente) e genera l'Excel.

Avvio:
    python demo_import.py
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook

from cronocantieri.analisi import analizza_cantiere
from cronocantieri.esporta_excel import esporta
from cronocantieri.models import Archivio
from cronocantieri.unisci import unisci_cronoprogrammi

CARTELLA = Path("import_imprese")


def crea_file_impresa(nome_file, meta, attivita):
    """Crea un Excel con metadati + tabella attività."""
    wb = Workbook()
    ws = wb.active
    for i, (et, val) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=et)
        ws.cell(row=i, column=2, value=val)
    r0 = len(meta) + 2
    for c, testo in enumerate(["Attività", "Data inizio", "Data fine",
                               "Importo", "Avanzamento %"], start=1):
        ws.cell(row=r0, column=c, value=testo)
    for i, riga in enumerate(attivita, start=1):
        for c, val in enumerate(riga, start=1):
            ws.cell(row=r0 + i, column=c, value=val)
    CARTELLA.mkdir(exist_ok=True)
    percorso = CARTELLA / nome_file
    wb.save(percorso)
    return percorso


def main() -> None:
    # --- Impresa 1: edile, con % dichiarata per attività
    f1 = crea_file_impresa(
        "impresa_rossi.xlsx",
        [("Cantiere:", "Scuola Media Via Roma"),
         ("Impresa:", "Rossi Costruzioni S.r.l."),
         ("Lotto:", "Lotto 1 - Opere edili"),
         ("Categoria:", "edile")],
        [("Scavi e fondazioni", "2026-01-07", "2026-02-28", 150000, 100),
         ("Struttura in c.a.", "2026-03-01", "2026-05-31", 500000, 60),
         ("Murature e tamponamenti", "2026-06-01", "2026-07-31", 200000, 0)],
    )

    # --- Impresa 2: impianti, SENZA % (verrà stimata dalle date)
    f2 = crea_file_impresa(
        "impresa_bianchi.xlsx",
        [("Cantiere:", "Scuola Media Via Roma"),
         ("Impresa:", "Bianchi Impianti S.p.A."),
         ("Lotto:", "Lotto 2 - Impianti"),
         ("Categoria:", "impiantistico")],
        [("Impianto elettrico", "2026-03-01", "2026-08-31", 350000, None),
         ("Impianto meccanico", "2026-04-01", "2026-09-15", 270000, None)],
    )

    # --- Impresa 3: serramenti, importi in formato italiano e % mista
    f3 = crea_file_impresa(
        "impresa_verde.xlsx",
        [("Cantiere:", "Scuola Media Via Roma"),
         ("Impresa:", "Verde Infissi S.r.l."),
         ("Lotto:", "Lotto 3 - Serramenti"),
         ("Categoria:", "serramenti")],
        [("Fornitura infissi", "2026-04-01", "2026-06-30", "180.000,00", 70),
         ("Posa in opera", "2026-07-01", "2026-08-31", "120.000,00", 10)],
    )

    print("File di esempio creati in:", CARTELLA.resolve())
    print("  -", f1.name, "/", f2.name, "/", f3.name)

    # --- Import + fusione
    cantiere, importati = unisci_cronoprogrammi(
        [f1, f2, f3],
        committente="Comune di Esempio",
    )
    print("\nFusione completata. Cantiere:", cantiere.nome)
    for imp in importati:
        print(f"  {imp.file_origine}: {imp.lotto.impresa} "
              f"({imp.righe_lette} attività, € {imp.lotto.importo_contratto:,.0f})")

    # salva nell'archivio
    archivio = Archivio()
    archivio.aggiungi_cantiere(cantiere)
    archivio.salva()

    # --- Analisi automatica al 1 giugno 2026
    stato = analizza_cantiere(cantiere, date(2026, 6, 1))
    print(f"\nAnalisi al 01/06/2026 - avanzamento globale "
          f"{stato.avanzamento_globale_reale}% "
          f"(atteso {stato.avanzamento_globale_atteso}%)")
    for s in stato.stati_lotti:
        print(f"  {s.lotto.impresa:<26} "
              f"calcolato {s.avanzamento_reale:>5}% / atteso {s.avanzamento_atteso:>5}%  "
              f"[{s.giudizio}]")
    peggiore = stato.lotto_piu_in_ritardo()
    if peggiore and peggiore.in_ritardo:
        print(f"\n>> Impresa piu in ritardo: {peggiore.lotto.impresa} "
              f"({peggiore.scostamento:+}%)")

    percorso = esporta(archivio, alla_data=date(2026, 6, 1))
    print(f"\nCronoprogramma totale esportato in: {percorso.resolve()}")


if __name__ == "__main__":
    main()
